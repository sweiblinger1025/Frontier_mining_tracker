"""
Export Dialog - Export various reports to Excel/CSV
"""
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QCheckBox, QRadioButton, QButtonGroup,
    QFileDialog, QMessageBox, QDateEdit, QProgressBar,
    QTabWidget, QWidget, QFormLayout, QComboBox
)
from PyQt6.QtCore import Qt, QDate
from datetime import datetime, date
import csv
import os


class ExportDialog(QDialog):
    """Dialog for exporting reports to Excel/CSV."""
    
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self.setWindowTitle("📤 Export Reports")
        self.setMinimumWidth(500)
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Tab widget for different export types
        tabs = QTabWidget()
        
        # Ledger Export Tab
        ledger_tab = QWidget()
        self._setup_ledger_tab(ledger_tab)
        tabs.addTab(ledger_tab, "📒 Ledger")
        
        # Summary Report Tab
        summary_tab = QWidget()
        self._setup_summary_tab(summary_tab)
        tabs.addTab(summary_tab, "📊 Summary")
        
        # Inventory Tab
        inventory_tab = QWidget()
        self._setup_inventory_tab(inventory_tab)
        tabs.addTab(inventory_tab, "📦 Inventory")
        
        # ROI Tab
        roi_tab = QWidget()
        self._setup_roi_tab(roi_tab)
        tabs.addTab(roi_tab, "💰 ROI")
        
        # Full Export Tab
        full_tab = QWidget()
        self._setup_full_tab(full_tab)
        tabs.addTab(full_tab, "📋 Full Report")
        
        layout.addWidget(tabs)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)
    
    def _setup_ledger_tab(self, parent):
        """Setup Ledger export options."""
        layout = QVBoxLayout(parent)
        
        # Description
        desc = QLabel("Export all Ledger transactions to a spreadsheet file.")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Date range filter
        range_group = QGroupBox("Date Range")
        range_layout = QFormLayout(range_group)
        
        self.ledger_all_dates = QCheckBox("All dates")
        self.ledger_all_dates.setChecked(True)
        self.ledger_all_dates.stateChanged.connect(self._on_ledger_dates_changed)
        range_layout.addRow(self.ledger_all_dates)
        
        self.ledger_from_date = QDateEdit()
        self.ledger_from_date.setDate(QDate(2021, 4, 22))
        self.ledger_from_date.setCalendarPopup(True)
        self.ledger_from_date.setEnabled(False)
        range_layout.addRow("From:", self.ledger_from_date)
        
        self.ledger_to_date = QDateEdit()
        self.ledger_to_date.setDate(QDate.currentDate())
        self.ledger_to_date.setCalendarPopup(True)
        self.ledger_to_date.setEnabled(False)
        range_layout.addRow("To:", self.ledger_to_date)
        
        layout.addWidget(range_group)
        
        # Format selection
        format_group = QGroupBox("Format")
        format_layout = QHBoxLayout(format_group)
        
        self.ledger_csv = QRadioButton("CSV")
        self.ledger_csv.setChecked(True)
        self.ledger_xlsx = QRadioButton("Excel (.xlsx)")
        
        format_layout.addWidget(self.ledger_csv)
        format_layout.addWidget(self.ledger_xlsx)
        format_layout.addStretch()
        
        layout.addWidget(format_group)
        
        # Export button
        export_btn = QPushButton("📤 Export Ledger")
        export_btn.clicked.connect(self._export_ledger)
        layout.addWidget(export_btn)
        
        layout.addStretch()
    
    def _on_ledger_dates_changed(self, state):
        """Toggle date range fields."""
        enabled = state != 2  # Not checked
        self.ledger_from_date.setEnabled(enabled)
        self.ledger_to_date.setEnabled(enabled)
    
    def _setup_summary_tab(self, parent):
        """Setup Summary report options."""
        layout = QVBoxLayout(parent)
        
        desc = QLabel("Generate a summary report with income, expenses, and net profit by period.")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Period selection
        period_group = QGroupBox("Report Period")
        period_layout = QFormLayout(period_group)
        
        self.summary_period = QComboBox()
        self.summary_period.addItems([
            "Daily (by date)",
            "Weekly",
            "Monthly",
            "All Time Total"
        ])
        period_layout.addRow("Group by:", self.summary_period)
        
        self.summary_from = QDateEdit()
        self.summary_from.setDate(QDate(2021, 4, 22))
        self.summary_from.setCalendarPopup(True)
        period_layout.addRow("From:", self.summary_from)
        
        self.summary_to = QDateEdit()
        self.summary_to.setDate(self._get_game_date())
        self.summary_to.setCalendarPopup(True)
        period_layout.addRow("To:", self.summary_to)
        
        layout.addWidget(period_group)
        
        # Export button
        export_btn = QPushButton("📤 Export Summary")
        export_btn.clicked.connect(self._export_summary)
        layout.addWidget(export_btn)
        
        layout.addStretch()
    
    def _setup_inventory_tab(self, parent):
        """Setup Inventory export options."""
        layout = QVBoxLayout(parent)
        
        desc = QLabel("Export current inventory levels with values and locations.")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Options
        options_group = QGroupBox("Include")
        options_layout = QVBoxLayout(options_group)
        
        self.inv_include_zero = QCheckBox("Include zero-quantity items")
        options_layout.addWidget(self.inv_include_zero)
        
        self.inv_include_value = QCheckBox("Include total values")
        self.inv_include_value.setChecked(True)
        options_layout.addWidget(self.inv_include_value)
        
        layout.addWidget(options_group)
        
        # Export button
        export_btn = QPushButton("📤 Export Inventory")
        export_btn.clicked.connect(self._export_inventory)
        layout.addWidget(export_btn)
        
        layout.addStretch()
    
    def _setup_roi_tab(self, parent):
        """Setup ROI export options."""
        layout = QVBoxLayout(parent)
        
        desc = QLabel("Export investment tracking data with ROI calculations.")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Options
        options_group = QGroupBox("Include")
        options_layout = QVBoxLayout(options_group)
        
        self.roi_include_revenue = QCheckBox("Include revenue history")
        self.roi_include_revenue.setChecked(True)
        options_layout.addWidget(self.roi_include_revenue)
        
        self.roi_include_maintenance = QCheckBox("Include maintenance records")
        self.roi_include_maintenance.setChecked(True)
        options_layout.addWidget(self.roi_include_maintenance)
        
        self.roi_include_fuel = QCheckBox("Include fuel costs by vehicle")
        self.roi_include_fuel.setChecked(True)
        options_layout.addWidget(self.roi_include_fuel)
        
        layout.addWidget(options_group)
        
        # Export button
        export_btn = QPushButton("📤 Export ROI Data")
        export_btn.clicked.connect(self._export_roi)
        layout.addWidget(export_btn)
        
        layout.addStretch()
    
    def _setup_full_tab(self, parent):
        """Setup Full report export."""
        layout = QVBoxLayout(parent)
        
        desc = QLabel(
            "Generate a comprehensive report including:\n"
            "• All Ledger transactions\n"
            "• Financial summary\n"
            "• Inventory snapshot\n"
            "• ROI tracking data\n"
            "• Production log\n"
            "• Maintenance records"
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Info
        info = QLabel("This will create a multi-sheet Excel workbook with all data.")
        info.setStyleSheet("color: #666; font-style: italic;")
        layout.addWidget(info)
        
        # Export button
        export_btn = QPushButton("📤 Export Full Report")
        export_btn.setStyleSheet("background-color: #D5F5D5; padding: 10px;")
        export_btn.clicked.connect(self._export_full)
        layout.addWidget(export_btn)
        
        layout.addStretch()
    
    def _get_game_date(self):
        """Get current game date."""
        try:
            if hasattr(self.main_window, 'settings_tab'):
                settings = self.main_window.settings_tab
                if hasattr(settings, 'current_game_date'):
                    d = settings.current_game_date.date()
                    return QDate(d.year(), d.month(), d.day())
        except:
            pass
        return QDate(2021, 4, 23)
    
    def _export_ledger(self):
        """Export ledger transactions."""
        # Get transactions
        if not hasattr(self.main_window, 'ledger_tab'):
            QMessageBox.warning(self, "Error", "Ledger not available")
            return
        
        transactions = self.main_window.ledger_tab.transactions
        if not transactions:
            QMessageBox.warning(self, "No Data", "No transactions to export")
            return
        
        # Filter by date if needed
        if not self.ledger_all_dates.isChecked():
            from_date = self.ledger_from_date.date().toString("yyyy-MM-dd")
            to_date = self.ledger_to_date.date().toString("yyyy-MM-dd")
            transactions = [
                t for t in transactions 
                if from_date <= str(t.get('date', ''))[:10] <= to_date
            ]
        
        # Get file path
        ext = "csv" if self.ledger_csv.isChecked() else "xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Ledger", f"ledger_export.{ext}",
            f"{'CSV' if ext == 'csv' else 'Excel'} Files (*.{ext})"
        )
        
        if not file_path:
            return
        
        try:
            if ext == "csv":
                self._write_ledger_csv(file_path, transactions)
            else:
                self._write_ledger_xlsx(file_path, transactions)
            
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {len(transactions)} transactions to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
    
    def _write_ledger_csv(self, path, transactions):
        """Write transactions to CSV."""
        headers = ['Date', 'Type', 'Item', 'Category', 'Quantity', 'Unit Price', 
                   'Subtotal', 'Discount', 'Total', 'Account', 'Location', 'Vehicle', 'Notes']
        
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for t in transactions:
                writer.writerow([
                    t.get('date', ''),
                    t.get('type', ''),
                    t.get('item', ''),
                    t.get('category', ''),
                    t.get('quantity', ''),
                    t.get('unit_price', ''),
                    t.get('subtotal', ''),
                    t.get('discount', ''),
                    t.get('total', ''),
                    t.get('account', ''),
                    t.get('location', ''),
                    t.get('vehicle', ''),
                    t.get('notes', ''),
                ])
    
    def _write_ledger_xlsx(self, path, transactions):
        """Write transactions to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Fallback to CSV
            csv_path = path.replace('.xlsx', '.csv')
            self._write_ledger_csv(csv_path, transactions)
            QMessageBox.information(
                self, "Note", 
                f"openpyxl not installed. Exported as CSV instead:\n{csv_path}"
            )
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ledger"
        
        # Headers
        headers = ['Date', 'Type', 'Item', 'Category', 'Quantity', 'Unit Price', 
                   'Subtotal', 'Discount', 'Total', 'Account', 'Location', 'Vehicle', 'Notes']
        
        header_fill = PatternFill(start_color="4A90A4", end_color="4A90A4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row, t in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=t.get('date', ''))
            ws.cell(row=row, column=2, value=t.get('type', ''))
            ws.cell(row=row, column=3, value=t.get('item', ''))
            ws.cell(row=row, column=4, value=t.get('category', ''))
            ws.cell(row=row, column=5, value=t.get('quantity', 0))
            ws.cell(row=row, column=6, value=t.get('unit_price', 0))
            ws.cell(row=row, column=7, value=t.get('subtotal', 0))
            ws.cell(row=row, column=8, value=t.get('discount', 0))
            ws.cell(row=row, column=9, value=t.get('total', 0))
            ws.cell(row=row, column=10, value=t.get('account', ''))
            ws.cell(row=row, column=11, value=t.get('location', ''))
            ws.cell(row=row, column=12, value=t.get('vehicle', ''))
            ws.cell(row=row, column=13, value=t.get('notes', ''))
        
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        wb.save(path)
    
    def _export_summary(self):
        """Export financial summary."""
        if not hasattr(self.main_window, 'ledger_tab'):
            QMessageBox.warning(self, "Error", "Ledger not available")
            return
        
        transactions = self.main_window.ledger_tab.transactions
        
        # Filter by date range
        from_date = self.summary_from.date().toString("yyyy-MM-dd")
        to_date = self.summary_to.date().toString("yyyy-MM-dd")
        
        filtered = [
            t for t in transactions 
            if from_date <= str(t.get('date', ''))[:10] <= to_date
            and t.get('type') != 'Opening'
        ]
        
        # Group by period
        period = self.summary_period.currentText()
        summary = {}
        
        for t in filtered:
            date_str = str(t.get('date', ''))[:10]
            
            if "Daily" in period:
                key = date_str
            elif "Weekly" in period:
                # Get week start
                d = datetime.strptime(date_str, "%Y-%m-%d")
                week_start = d - timedelta(days=d.weekday())
                key = week_start.strftime("%Y-%m-%d")
            elif "Monthly" in period:
                key = date_str[:7]  # YYYY-MM
            else:
                key = "Total"
            
            if key not in summary:
                summary[key] = {'income': 0, 'expenses': 0}
            
            total = t.get('total', 0)
            if t.get('type') == 'Sale':
                summary[key]['income'] += abs(total)
            elif t.get('type') in ['Purchase', 'Fuel']:
                summary[key]['expenses'] += abs(total)
        
        # Export
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Summary", "financial_summary.csv",
            "CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            from datetime import timedelta
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Period', 'Income', 'Expenses', 'Net'])
                
                for period_key in sorted(summary.keys()):
                    data = summary[period_key]
                    net = data['income'] - data['expenses']
                    writer.writerow([
                        period_key,
                        f"${data['income']:,.0f}",
                        f"${data['expenses']:,.0f}",
                        f"${net:,.0f}"
                    ])
                
                # Total row
                total_income = sum(d['income'] for d in summary.values())
                total_expenses = sum(d['expenses'] for d in summary.values())
                writer.writerow([])
                writer.writerow([
                    'TOTAL',
                    f"${total_income:,.0f}",
                    f"${total_expenses:,.0f}",
                    f"${total_income - total_expenses:,.0f}"
                ])
            
            QMessageBox.information(
                self, "Export Complete",
                f"Summary exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
    
    def _export_inventory(self):
        """Export inventory data."""
        if not hasattr(self.main_window, 'inventory_tab'):
            QMessageBox.warning(self, "Error", "Inventory not available")
            return
        
        items = self.main_window.inventory_tab.inventory_items
        
        if not self.inv_include_zero.isChecked():
            items = [i for i in items if i.get('quantity', 0) > 0]
        
        if not items:
            QMessageBox.warning(self, "No Data", "No inventory items to export")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Inventory", "inventory_export.csv",
            "CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                headers = ['Item', 'Category', 'Location', 'Quantity', 'Unit Price']
                if self.inv_include_value.isChecked():
                    headers.append('Total Value')
                writer.writerow(headers)
                
                total_value = 0
                for item in items:
                    qty = item.get('quantity', 0)
                    price = item.get('unit_price', 0)
                    value = qty * price
                    total_value += value
                    
                    row = [
                        item.get('name', ''),
                        item.get('category', ''),
                        item.get('location', ''),
                        qty,
                        f"${price:,.2f}"
                    ]
                    if self.inv_include_value.isChecked():
                        row.append(f"${value:,.2f}")
                    writer.writerow(row)
                
                if self.inv_include_value.isChecked():
                    writer.writerow([])
                    writer.writerow(['', '', '', 'TOTAL:', '', f"${total_value:,.2f}"])
            
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {len(items)} items to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
    
    def _export_roi(self):
        """Export ROI tracking data."""
        if not hasattr(self.main_window, 'roi_tracker_tab'):
            QMessageBox.warning(self, "Error", "ROI Tracker not available")
            return
        
        roi_tab = self.main_window.roi_tracker_tab
        investments = roi_tab.investments
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export ROI Data", "roi_export.csv",
            "CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Investments section
                writer.writerow(['=== INVESTMENTS ==='])
                writer.writerow(['Name', 'Category', 'Cost', 'Revenue', 'Profit', 'ROI %', 'Purchase Date'])
                
                for inv in investments:
                    revenue = sum(r.get('amount', 0) for r in inv.get('revenues', []))
                    cost = inv.get('cost', 0)
                    profit = revenue - cost
                    roi = (profit / cost * 100) if cost > 0 else 0
                    
                    writer.writerow([
                        inv.get('name', ''),
                        inv.get('category', ''),
                        f"${cost:,.0f}",
                        f"${revenue:,.0f}",
                        f"${profit:,.0f}",
                        f"{roi:.1f}%",
                        inv.get('purchase_date', '')
                    ])
                
                # Maintenance section
                if self.roi_include_maintenance.isChecked() and hasattr(roi_tab, 'maintenance_records'):
                    writer.writerow([])
                    writer.writerow(['=== MAINTENANCE ==='])
                    writer.writerow(['Date', 'Equipment', 'Type', 'Cost', 'Notes'])
                    
                    for rec in roi_tab.maintenance_records:
                        writer.writerow([
                            rec.get('date', ''),
                            rec.get('equipment', ''),
                            rec.get('type', ''),
                            f"${rec.get('cost', 0):,.0f}",
                            rec.get('notes', '')
                        ])
                
                # Fuel section
                if self.roi_include_fuel.isChecked():
                    writer.writerow([])
                    writer.writerow(['=== FUEL COSTS ==='])
                    writer.writerow(['Vehicle', 'Liters', 'Cost', 'Transactions'])
                    
                    if hasattr(self.main_window, 'ledger_tab'):
                        fuel_data = self.main_window.ledger_tab.get_fuel_by_vehicle()
                        for vehicle, data in sorted(fuel_data.items()):
                            writer.writerow([
                                vehicle,
                                data['liters'],
                                f"${data['cost']:,.0f}",
                                data['transactions']
                            ])
            
            QMessageBox.information(
                self, "Export Complete",
                f"ROI data exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
    
    def _export_full(self):
        """Export comprehensive report."""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Full Report", "frontier_mining_full_report.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            QMessageBox.warning(
                self, "Missing Dependency",
                "Full report export requires openpyxl.\n\n"
                "Install with: pip install openpyxl"
            )
            return
        
        try:
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            header_fill = PatternFill(start_color="4A90A4", end_color="4A90A4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Sheet 1: Ledger
            if hasattr(self.main_window, 'ledger_tab'):
                ws = wb.create_sheet("Ledger")
                transactions = self.main_window.ledger_tab.transactions
                
                headers = ['Date', 'Type', 'Item', 'Category', 'Qty', 'Unit Price', 
                           'Total', 'Account', 'Notes']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row, t in enumerate(transactions, 2):
                    ws.cell(row=row, column=1, value=str(t.get('date', ''))[:10])
                    ws.cell(row=row, column=2, value=t.get('type', ''))
                    ws.cell(row=row, column=3, value=t.get('item', ''))
                    ws.cell(row=row, column=4, value=t.get('category', ''))
                    ws.cell(row=row, column=5, value=t.get('quantity', 0))
                    ws.cell(row=row, column=6, value=t.get('unit_price', 0))
                    ws.cell(row=row, column=7, value=t.get('total', 0))
                    ws.cell(row=row, column=8, value=t.get('account', ''))
                    ws.cell(row=row, column=9, value=t.get('notes', ''))
            
            # Sheet 2: Inventory
            if hasattr(self.main_window, 'inventory_tab'):
                ws = wb.create_sheet("Inventory")
                items = self.main_window.inventory_tab.inventory_items
                
                headers = ['Item', 'Category', 'Location', 'Quantity', 'Unit Price', 'Value']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row, item in enumerate(items, 2):
                    qty = item.get('quantity', 0)
                    price = item.get('unit_price', 0)
                    ws.cell(row=row, column=1, value=item.get('name', ''))
                    ws.cell(row=row, column=2, value=item.get('category', ''))
                    ws.cell(row=row, column=3, value=item.get('location', ''))
                    ws.cell(row=row, column=4, value=qty)
                    ws.cell(row=row, column=5, value=price)
                    ws.cell(row=row, column=6, value=qty * price)
            
            # Sheet 3: Investments
            if hasattr(self.main_window, 'roi_tracker_tab'):
                ws = wb.create_sheet("Investments")
                investments = self.main_window.roi_tracker_tab.investments
                
                headers = ['Name', 'Category', 'Cost', 'Revenue', 'Profit', 'ROI %']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row, inv in enumerate(investments, 2):
                    revenue = sum(r.get('amount', 0) for r in inv.get('revenues', []))
                    cost = inv.get('cost', 0)
                    profit = revenue - cost
                    roi = (profit / cost * 100) if cost > 0 else 0
                    
                    ws.cell(row=row, column=1, value=inv.get('name', ''))
                    ws.cell(row=row, column=2, value=inv.get('category', ''))
                    ws.cell(row=row, column=3, value=cost)
                    ws.cell(row=row, column=4, value=revenue)
                    ws.cell(row=row, column=5, value=profit)
                    ws.cell(row=row, column=6, value=f"{roi:.1f}%")
            
            # Sheet 4: Production
            if hasattr(self.main_window, 'production_tab'):
                ws = wb.create_sheet("Production")
                log = self.main_window.production_tab.production_log
                
                headers = ['Date', 'Building', 'Output', 'Quantity', 'Value Created']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                
                for row, entry in enumerate(log, 2):
                    dt = entry.get('datetime')
                    ws.cell(row=row, column=1, value=dt.strftime("%Y-%m-%d %H:%M") if dt else '')
                    ws.cell(row=row, column=2, value=entry.get('building', ''))
                    ws.cell(row=row, column=3, value=entry.get('output', ''))
                    ws.cell(row=row, column=4, value=entry.get('output_qty', 0))
                    ws.cell(row=row, column=5, value=entry.get('value_created', 0))
            
            # Sheet 5: Maintenance
            if hasattr(self.main_window, 'roi_tracker_tab'):
                roi_tab = self.main_window.roi_tracker_tab
                if hasattr(roi_tab, 'maintenance_records') and roi_tab.maintenance_records:
                    ws = wb.create_sheet("Maintenance")
                    
                    headers = ['Date', 'Equipment', 'Type', 'Cost', 'Notes']
                    for col, h in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row, rec in enumerate(roi_tab.maintenance_records, 2):
                        ws.cell(row=row, column=1, value=rec.get('date', ''))
                        ws.cell(row=row, column=2, value=rec.get('equipment', ''))
                        ws.cell(row=row, column=3, value=rec.get('type', ''))
                        ws.cell(row=row, column=4, value=rec.get('cost', 0))
                        ws.cell(row=row, column=5, value=rec.get('notes', ''))
            
            wb.save(file_path)
            
            QMessageBox.information(
                self, "Export Complete",
                f"Full report exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
